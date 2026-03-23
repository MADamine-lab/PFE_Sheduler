"""
scheduler/export.py

Excel and PDF export views.

Flask equivalent: routes/export.py  (export_bp Blueprint)

Key differences vs Flask
──────────────────────────────────────────────────────────────────────────────
| Flask                              | Django / DRF                           |
|──────────────────────────────────────────────────────────────────────────── |
| send_file(buf, mimetype=...,       | FileResponse(buf, content_type=...)    |
|           as_attachment=True,      | with Content-Disposition header set    |
|           download_name="x.xlsx") | manually                               |
|──────────────────────────────────────────────────────────────────────────── |
| Affectation.query.all()            | Affectation.objects.select_related()   |
|                                    | (avoids N+1 queries)                   |
|──────────────────────────────────────────────────────────────────────────── |
| return jsonify({"error":...}), 501 | return Response({...},                 |
|                                    |   status=HTTP_501_NOT_IMPLEMENTED)     |
──────────────────────────────────────────────────────────────────────────────

The Excel and PDF generation logic (openpyxl / reportlab) is completely
unchanged from your Flask version — only the view wrapper changes.
"""

import io
from django.http import FileResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

from .models import Affectation


# ── Shared helper (unchanged from Flask version) ──────────────────────────────

def _style_header(ws, row, ncols, color="1F4E79"):
    """
    Applies bold white text on a dark background to a header row.
    Identical to the Flask version — pure openpyxl, no Flask dependency.
    """
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
#  Flask: @export_bp.route("/excel", methods=["GET"])
#  Django: GET /api/export/excel/
# ══════════════════════════════════════════════════════════════════════════════

class ExportExcelView(APIView):
    """
    Generates and streams a multi-sheet Excel workbook.

    Sheet 1 — "Planning Soutenances" : all assignments
    Sheet 2…N — one sheet per domain (Informatique, Electrique, …)

    Flask used send_file(buf, as_attachment=True, download_name="x.xlsx").
    Django uses FileResponse which streams the BytesIO buffer directly.
    The Content-Disposition header tells the browser to download the file.
    """

    def get(self, request):

        # select_related → single SQL JOIN for all FK relations
        # Flask: Affectation.query.all()  (caused N+1 queries)
        # Django: select_related fetches everything in one query
        affectations = Affectation.objects.select_related(
            "etudiant__encadrant",
            "examinateur",
            "president",
            "creneau",
        ).all()

        wb = openpyxl.Workbook()

        # ── Sheet 1: Full planning ─────────────────────────────────────────────
        ws = wb.active
        ws.title = "Planning Soutenances"

        headers = [
            "Étudiant ID", "Nom", "Prénom", "Domaine", "Sujet",
            "Encadrant", "Examinateur", "Président",
            "Date", "Créneau", "Salle",
            "Score Exam.", "Score Prés.",
        ]
        ws.append(headers)
        _style_header(ws, 1, len(headers))
        ws.row_dimensions[1].height = 30

        domain_colors = {
            "Informatique": "DDEEFF",
            "Electrique":   "DDFFEE",
            "Mecanique":    "FFF3CD",
            "Energetique":  "FFE0E0",
            "Genie Civil":  "F0E6FF",
        }

        for i, aff in enumerate(affectations, 2):
            # Django ORM: access related objects via dot notation
            # select_related() means these do NOT trigger extra DB queries
            etu  = aff.etudiant
            enc  = etu.encadrant if etu else None    # etu.encadrant → Professeur
            exam = aff.examinateur
            pres = aff.president
            cr   = aff.creneau

            row = [
                etu.id     if etu  else "",
                etu.nom    if etu  else "",
                etu.prenom if etu  else "",
                etu.domaine if etu else "",
                etu.sujet  if etu  else "",
                f"{enc.prenom} {enc.nom}"   if enc  else "",
                f"{exam.prenom} {exam.nom}" if exam else "",
                f"{pres.prenom} {pres.nom}" if pres else "",
                cr.date  if cr else "",
                cr.slot  if cr else "",
                cr.salle if cr else "",
                round(aff.score_exam, 2),
                round(aff.score_pres, 2),
            ]
            ws.append(row)

            color = domain_colors.get(etu.domaine if etu else "", "F5F5F5")
            fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill
                ws.cell(row=i, column=col).alignment = Alignment(
                    wrap_text=True, vertical="center"
                )

        col_widths = [12, 14, 12, 14, 60, 22, 22, 22, 13, 16, 18, 13, 13]
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

        # ── Sheets 2…N: One per domain ─────────────────────────────────────────
        by_domain = defaultdict(list)
        for aff in affectations:
            dom = aff.etudiant.domaine if aff.etudiant else "Autre"
            by_domain[dom].append(aff)

        for dom, affs in by_domain.items():
            ws2 = wb.create_sheet(dom[:25])
            ws2.append(headers)
            _style_header(ws2, 1, len(headers), "145A32")

            for i, aff in enumerate(affs, 2):
                etu  = aff.etudiant
                enc  = etu.encadrant if etu else None
                exam = aff.examinateur
                pres = aff.president
                cr   = aff.creneau

                ws2.append([
                    etu.id     if etu  else "",
                    etu.nom    if etu  else "",
                    etu.prenom if etu  else "",
                    etu.domaine if etu else "",
                    etu.sujet  if etu  else "",
                    f"{enc.prenom} {enc.nom}"   if enc  else "",
                    f"{exam.prenom} {exam.nom}" if exam else "",
                    f"{pres.prenom} {pres.nom}" if pres else "",
                    cr.date  if cr else "",
                    cr.slot  if cr else "",
                    cr.salle if cr else "",
                    round(aff.score_exam, 2),
                    round(aff.score_pres, 2),
                ])

                color = "EAF3DE" if i % 2 == 0 else "FFFFFF"
                fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws2.cell(row=i, column=col).fill = fill

            for j, w in enumerate(col_widths, 1):
                ws2.column_dimensions[get_column_letter(j)].width = w

        # ── Stream workbook to client ──────────────────────────────────────────
        # Flask:  send_file(buf, mimetype="application/vnd...", as_attachment=True,
        #                   download_name="planning_soutenances_PFE.xlsx")
        #
        # Django: FileResponse streams the buffer.
        #         Content-Disposition tells the browser to download the file.
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = FileResponse(
            buf,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="planning_soutenances_PFE.xlsx"'
        )
        return response


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORT
#  Flask: @export_bp.route("/pdf", methods=["GET"])
#  Django: GET /api/export/pdf/
# ══════════════════════════════════════════════════════════════════════════════

class ExportPdfView(APIView):
    """
    Generates and streams a landscape A4 PDF using reportlab.

    The reportlab generation code is 100% identical to the Flask version.
    Only the response wrapper changes:
        Flask  → send_file(buf, mimetype="application/pdf", as_attachment=True)
        Django → FileResponse(buf, content_type="application/pdf")
    """

    def get(self, request):

        # ── Check reportlab is installed ───────────────────────────────────────
        # Flask: try/except ImportError → return jsonify({"error":...}), 501
        # Django: same pattern, but return DRF Response with HTTP 501
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            return Response(
                {"error": "reportlab non installé — utilisez l'export Excel"},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

        # Fetch all assignments with related objects in one query
        affectations = Affectation.objects.select_related(
            "etudiant__encadrant",
            "examinateur",
            "president",
            "creneau",
        ).all()

        # ── Build PDF (identical to Flask version) ─────────────────────────────
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=1*cm, rightMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm,
        )

        styles      = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "title", parent=styles["Heading1"], fontSize=14,
            spaceAfter=12, textColor=colors.HexColor("#1F4E79"),
        )
        small = ParagraphStyle(
            "small", parent=styles["Normal"], fontSize=7, leading=9
        )

        story = [
            Paragraph("Planning des Soutenances PFE", title_style),
            Spacer(1, 0.3*cm),
        ]

        headers = [
            "Étudiant", "Domaine", "Sujet",
            "Encadrant", "Examinateur", "Président",
            "Date", "Salle",
        ]
        data = [headers]

        for aff in affectations:
            etu  = aff.etudiant
            enc  = etu.encadrant if etu else None
            exam = aff.examinateur
            pres = aff.president
            cr   = aff.creneau

            sujet = ""
            if etu:
                sujet = (etu.sujet[:55] + "…") if len(etu.sujet) > 55 else etu.sujet

            data.append([
                Paragraph(f"{etu.prenom} {etu.nom}" if etu else "", small),
                Paragraph(etu.domaine[:12]          if etu else "", small),
                Paragraph(sujet,                                     small),
                Paragraph(enc.nom  if enc  else "", small),
                Paragraph(exam.nom if exam else "", small),
                Paragraph(pres.nom if pres else "", small),
                Paragraph(f"{cr.date}\n{cr.slot}" if cr else "", small),
                Paragraph(cr.salle if cr else "",                 small),
            ])

        col_widths_pdf = [
            3.2*cm, 2.2*cm, 7.5*cm,
            2.8*cm, 2.8*cm, 2.8*cm,
            2.6*cm, 2.5*cm,
        ]

        t = Table(data, colWidths=col_widths_pdf, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#1F4E79")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0),  8),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ]))

        story.append(t)
        doc.build(story)
        buf.seek(0)

        # ── Stream PDF to client ───────────────────────────────────────────────
        # Flask:  send_file(buf, mimetype="application/pdf",
        #                   as_attachment=True, download_name="planning_PFE.pdf")
        # Django: FileResponse with Content-Disposition header
        response = FileResponse(buf, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="planning_PFE.pdf"'
        return response